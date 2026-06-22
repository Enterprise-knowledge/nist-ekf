#!/usr/bin/env python3
"""Generate the NIST EKF demo bundle from local PDFs and NIST data exports."""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pdfplumber
import requests
import yaml
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SKILL = ROOT / ".agents" / "skills" / "ekf-bootstrap"
UPDATED = "2026-06-22T00:00:00Z"
OWNER = {"name": "Enterprise Knowledge Format Maintainers"}

URLS = {
    "csf_excel": "https://csrc.nist.gov/extensions/nudp/services/json/nudp/framework/version/CSF_2_0_0/export/excel",
    "csf_functions": "https://csrc.nist.gov/extensions/nudp/services/json/nudp/framework/version/CSF_2_0_0/type/function/elements",
    "csf_categories": "https://csrc.nist.gov/extensions/nudp/services/json/nudp/framework/version/CSF_2_0_0/type/category/elements",
    "csf_subcategories": "https://csrc.nist.gov/extensions/nudp/services/json/nudp/framework/version/CSF_2_0_0/type/subcategory/elements",
    "csf_examples": "https://csrc.nist.gov/extensions/nudp/services/json/nudp/framework/version/CSF_2_0_0/type/implementation_example/elements",
    "ai_playbook_json": "https://airc.nist.gov/docs/playbook.json",
    "sp80053_oscal": "https://raw.githubusercontent.com/usnistgov/oscal-content/main/nist.gov/SP800-53/rev5/json/NIST_SP-800-53_rev5_catalog.json",
}

PDFS = {
    "csf": {
        "path": ROOT / "source" / "NIST.CSWP.29.pdf",
        "artifact": "/artifacts/markdown/documents/NIST.CSWP.29.md",
        "doi": "https://doi.org/10.6028/NIST.CSWP.29",
    },
    "ai": {
        "path": ROOT / "source" / "NIST.AI.100-1.pdf",
        "artifact": "/artifacts/markdown/documents/NIST.AI.100-1.md",
        "doi": "https://doi.org/10.6028/NIST.AI.100-1",
    },
    "sp80053": {
        "path": ROOT / "source" / "NIST.SP.800-53r5.pdf",
        "artifact": "/artifacts/markdown/documents/NIST.SP.800-53r5.md",
        "doi": "https://doi.org/10.6028/NIST.SP.800-53r5",
    },
}

CSF_ACTIVE_CATEGORIES = {
    "GV.OC",
    "GV.RM",
    "GV.RR",
    "GV.PO",
    "GV.OV",
    "GV.SC",
    "ID.AM",
    "ID.RA",
    "ID.IM",
    "PR.AA",
    "PR.AT",
    "PR.DS",
    "PR.PS",
    "PR.IR",
    "DE.CM",
    "DE.AE",
    "RS.MA",
    "RS.AN",
    "RS.CO",
    "RS.MI",
    "RC.RP",
    "RC.CO",
}


def slug(value: str) -> str:
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "item"


def mkdir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def clean_generated() -> None:
    for path in [
        ROOT / "bundles",
        ROOT / "artifacts" / "markdown",
        ROOT / "artifacts" / "data",
        ROOT / "artifacts" / "graph",
        ROOT / "artifacts" / "html",
        ROOT / "knowledge" / "domains" / "nist-governance-corpus.md",
        ROOT / "knowledge" / "systems" / "nist-reference-library.md",
        ROOT / "knowledge" / "processes" / "nist-risk-management-integration.md",
    ]:
        if path.is_dir():
            shutil.rmtree(path)
        elif path.exists():
            path.unlink()


def write_text(path: Path, text: str) -> None:
    mkdir(path.parent)
    path.write_text(text, encoding="utf-8")


def dump_yaml(data: Any) -> str:
    return yaml.safe_dump(data, sort_keys=False, allow_unicode=True, width=110).strip()


def md(frontmatter: dict[str, Any], body: str) -> str:
    return f"---\n{dump_yaml(frontmatter)}\n---\n\n{body.strip()}\n"


def link(path: str, title: str, description: str) -> str:
    return f"- [{title}]({path}) - {description}"


def source_ref(name: str, path: str, kind: str, description: str) -> dict[str, str]:
    return {"name": name, "path": path, "kind": kind, "description": description}


def relation(name: str, path: str, rel: str, description: str) -> dict[str, str]:
    return {"name": name, "path": path, "relation": rel, "description": description}


def extract_csf_active_subcategory_ids() -> set[str]:
    """Read Appendix A of the local CSF PDF for active CSF 2.0 Subcategory IDs."""
    ids: set[str] = set()
    with pdfplumber.open(str(PDFS["csf"]["path"])) as pdf:
        for page in pdf.pages[20:28]:
            text = page.extract_text() or ""
            ids.update(re.findall(r"(?m)^([A-Z]{2}\.[A-Z]{2}-\d{2}):", text))
    return ids


def base_fm(
    type_: str,
    title: str,
    description: str,
    tags: list[str],
    sources: list[dict[str, str]],
    related: list[dict[str, str]],
    domain: str,
    system: str,
    status: str = "active",
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    fm: dict[str, Any] = {
        "type": type_,
        "title": title,
        "description": description,
        "status": status,
        "owner": OWNER,
        "updated": UPDATED,
        "domain": domain,
        "system": system,
        "tags": tags,
        "confidence": "high",
        "review": {
            "status": "agent-generated",
            "reviewed_by": "EKF generation workflow",
            "reviewed_at": UPDATED,
        },
        "sources": sources,
        "related": related,
    }
    if extra:
        fm.update(extra)
    return fm


def index_fm(title: str, description: str, status: str = "active") -> dict[str, Any]:
    return {
        "type": "index",
        "title": title,
        "description": description,
        "status": status,
        "owner": OWNER,
        "updated": UPDATED,
    }


def bundle_fm(bundle_id: str, title: str, description: str, parent: str = "/") -> dict[str, Any]:
    return {
        "type": "bundle",
        "title": title,
        "description": description,
        "ekf_version": "0.1",
        "bundle": {"id": bundle_id, "kind": "nested", "parent": parent},
        "status": "active",
        "owner": OWNER,
        "updated": UPDATED,
    }


def download(url: str, path: Path) -> None:
    mkdir(path.parent)
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    path.write_bytes(response.content)


def extract_pdf_markdown(key: str, info: dict[str, Any]) -> None:
    pdf_path = info["path"]
    reader = PdfReader(str(pdf_path))
    meta = reader.metadata or {}
    lines = [
        f"# {meta.get('/Title') or pdf_path.stem}",
        "",
        f"- Source PDF: `/source/{pdf_path.name}`",
        f"- DOI: {info['doi']}",
        f"- Pages: {len(reader.pages)}",
        f"- Extracted: {datetime.now(UTC).replace(microsecond=0).isoformat()}",
        "",
    ]
    with pdfplumber.open(str(pdf_path)) as pdf:
        for number, page in enumerate(pdf.pages, start=1):
            lines.append(f"## Page {number}")
            lines.append("")
            lines.append(page.extract_text() or "")
            lines.append("")
    write_text(ROOT / info["artifact"].lstrip("/"), "\n".join(lines))


def concept_path(bundle: str, rel: str) -> str:
    return f"/bundles/{bundle}/knowledge/{rel}"


def local_concept_path(bundle: str, rel: str) -> Path:
    return ROOT / "bundles" / bundle / "knowledge" / rel


def create_bundle(bundle_id: str, title: str, description: str, subdirs: dict[str, str]) -> None:
    root = ROOT / "bundles" / bundle_id
    write_text(
        root / "index.md",
        md(
            bundle_fm(bundle_id, title, description),
            f"""# {title}

{description}

## Knowledge

- [Knowledge](knowledge/) - Canonical concepts extracted from the NIST publication and supplemental NIST data.

## Sources

- [Sources](source/) - Source provenance and references back to the root source corpus.

## Artifacts

- [Artifacts](artifacts/) - Bundle-local generated indexes and derived artifacts.
""",
        ),
    )
    write_text(root / "source" / "index.md", f"# {title} Sources\n\nThis bundle cites source material stored in the root `/source/` directory and generated artifacts under root `/artifacts/`.\n")
    write_text(root / "artifacts" / "index.md", f"# {title} Artifacts\n\nBundle-specific generated artifacts may be placed here. Shared graph and extraction artifacts are stored at the root.\n")
    write_text(root / "meta" / "conventions.md", f"# {title} Conventions\n\nGenerated concept identifiers preserve NIST framework identifiers where possible.\n")
    write_text(root / "meta" / "governance.md", f"# {title} Governance\n\nGenerated content should be reviewed against the cited NIST publication before policy adoption.\n")
    write_text(root / "meta" / "ingestion.md", f"# {title} Ingestion\n\nThis bundle was generated by `scripts/generate_nist_ekf.py` from local PDFs and official NIST machine-readable exports.\n")

    subbundle_lines = [link(f"{d}/", t, desc) for d, (t, desc) in subdirs.items()]
    write_text(
        root / "knowledge" / "index.md",
        md(
            index_fm(f"{title} Knowledge", f"Canonical EKF concepts for {title}."),
            f"""# {title} Knowledge

## Subbundles

{chr(10).join(subbundle_lines)}

## Core Concepts

- [Publication](publication.md) - Publication-level concept and provenance for this NIST source.
""",
        ),
    )
    for dirname, (sub_title, sub_desc) in subdirs.items():
        write_text(
            root / "knowledge" / dirname / "index.md",
            md(index_fm(sub_title, sub_desc), f"# {sub_title}\n\n{sub_desc}\n\n## Core Concepts\n\nGenerated concepts in this subbundle are listed by the parent publication indexes and are discoverable with `rg '^type:'`.\n"),
        )


def root_indexes() -> None:
    root_fm = {
        "type": "bundle",
        "title": "NIST Governance EKF Demo",
        "description": "Enterprise Knowledge Format demo bundle for NIST cybersecurity, AI risk, and security control publications.",
        "ekf_version": "0.1",
        "bundle": {"id": "nist-governance-ekf-demo", "kind": "root"},
        "status": "active",
        "owner": OWNER,
        "updated": UPDATED,
    }
    write_text(
        ROOT / "index.md",
        md(
            root_fm,
            """# NIST Governance EKF Demo

This EKF bundle is a publication-backed knowledge base for NIST governance material. It keeps local PDFs as source material, generated full-text markdown as artifacts, curated concepts under `knowledge/`, and publication-specific nested bundles under `bundles/`.

## Knowledge

- [Knowledge](knowledge/) - Cross-publication concepts and discovery indexes.

## Sources

- [Sources](source/) - Local NIST PDF source files and source manifest.

## Artifacts

- [Artifacts](artifacts/) - Generated markdown extractions, NIST data exports, graph JSON, and HTML graph demo.

## Nested Bundles

- [NIST Cybersecurity Framework 2.0](bundles/nist-csf-2/) - CSF 2.0 Functions, Categories, Subcategories, Profiles, Tiers, and Implementation Examples.
- [NIST AI RMF 1.0](bundles/nist-ai-rmf-1/) - AI RMF Core Functions, Categories, Subcategories, Playbook content, trustworthiness characteristics, and profiles.
- [NIST SP 800-53 Rev. 5](bundles/nist-sp-800-53-r5/) - Security and privacy control families, controls, control enhancements, publication sections, and glossary/acronym material.
""",
        ),
    )
    write_text(
        ROOT / "knowledge" / "index.md",
        md(
            index_fm("NIST Governance EKF Demo Knowledge", "Cross-publication discovery index for canonical EKF concepts."),
            """# NIST Governance EKF Demo Knowledge

## Subbundles

- [Domains](domains/) - Cross-publication governance domains.
- [Systems](systems/) - Reference systems and structured source systems used by this demo.
- [APIs](apis/) - API and machine-readable source concepts, if added later.
- [Data](data/) - Data exports and schema concepts, if added later.
- [Code](code/) - Generation scripts and implementation concepts, if added later.
- [Processes](processes/) - Risk management and ingestion processes.
- [Glossary](glossary/) - Shared terminology routed to publication-specific glossary bundles.

## Core Concepts

- [NIST Governance Corpus](domains/nist-governance-corpus.md) - Root concept connecting the three NIST publications.
- [NIST Reference Library](systems/nist-reference-library.md) - Source system concept for the local PDFs and official NIST exports.
- [NIST Risk Management Integration](processes/nist-risk-management-integration.md) - Cross-publication integration concept linking CSF, AI RMF, and SP 800-53 controls.
""",
        ),
    )
    root_subdirs = {
        "domains": (
            "Domains",
            "Cross-publication governance domains represented in this EKF demo.",
            [("nist-governance-corpus.md", "NIST Governance Corpus", "Cross-publication domain connecting cybersecurity risk, AI risk, and security and privacy controls.")],
        ),
        "systems": (
            "Systems",
            "Reference systems and structured source systems used by this demo.",
            [("nist-reference-library.md", "NIST Reference Library", "Source system for local NIST PDFs and official NIST exports.")],
        ),
        "apis": (
            "APIs",
            "API contracts, endpoints, events, and integration behavior.",
            [],
        ),
        "data": (
            "Data",
            "Generated data exports and structured data artifacts used by the demo.",
            [],
        ),
        "code": (
            "Code",
            "Generation scripts and implementation concepts for the EKF demo.",
            [],
        ),
        "processes": (
            "Processes",
            "Business processes, operations, runbooks, and decisions.",
            [("nist-risk-management-integration.md", "NIST Risk Management Integration", "Cross-publication integration concept linking CSF, AI RMF, and SP 800-53 controls.")],
        ),
        "glossary": (
            "Glossary",
            "Shared terminology for this bundle; publication-specific terms live in nested bundles.",
            [],
        ),
    }
    for dirname, (title, description, concepts) in root_subdirs.items():
        concept_lines = (
            "\n".join(link(path, concept_title, concept_desc) for path, concept_title, concept_desc in concepts)
            if concepts
            else "Publication-specific concepts are maintained in the nested bundles listed from the root index."
        )
        write_text(
            ROOT / "knowledge" / dirname / "index.md",
            md(
                index_fm(title, description),
                f"""# {title}

{description}

## Core Concepts

{concept_lines}
""",
            ),
        )


def source_and_artifact_indexes() -> None:
    manifest = {
        "id": "nist-governance-source-corpus",
        "name": "NIST Governance Source Corpus",
        "kind": "document",
        "description": "Local source PDFs for CSF 2.0, AI RMF 1.0, and SP 800-53 Rev. 5 with supplemental official NIST machine-readable exports.",
        "owner": OWNER,
        "location": "/source",
        "upstream": {"uri": "https://csrc.nist.gov/ and https://airc.nist.gov/", "ref": "NIST public publications and data exports"},
        "sync": {"mode": "manual", "last_synced": UPDATED},
        "indexing": {"include": ["*.pdf"], "exclude": []},
        "trust": {"level": "authoritative"},
    }
    write_text(ROOT / "source" / "ekf-source.yml", dump_yaml(manifest) + "\n")
    write_text(
        ROOT / "source" / "index.md",
        """# NIST Governance EKF Demo Sources

Raw NIST source publications are stored here.

## Source Files

- [NIST.CSWP.29.pdf](NIST.CSWP.29.pdf) - NIST Cybersecurity Framework (CSF) 2.0.
- [NIST.AI.100-1.pdf](NIST.AI.100-1.pdf) - Artificial Intelligence Risk Management Framework (AI RMF 1.0).
- [NIST.SP.800-53r5.pdf](NIST.SP.800-53r5.pdf) - Security and Privacy Controls for Information Systems and Organizations.
- [ekf-source.yml](ekf-source.yml) - Source manifest for indexing and provenance.
""",
    )
    write_text(
        ROOT / "artifacts" / "index.md",
        """# NIST Governance EKF Demo Artifacts

Generated representations belong here. Canonical EKF knowledge remains under `knowledge/` and nested bundle `knowledge/` directories.

## Artifact Areas

- [Markdown](markdown/) - Full page-marked markdown extractions of every source PDF.
- [Data](data/) - Downloaded official NIST machine-readable exports used for structured generation.
- [Graph](graph/) - Parsed EKF graph JSON.
- [HTML](html/) - Self-contained HTML knowledge graph demo.
""",
    )


def schemas() -> None:
    types = [
        "domain",
        "system",
        "service",
        "api",
        "api_endpoint",
        "database",
        "schema",
        "table",
        "codebase",
        "module",
        "business_process",
        "decision",
        "policy",
        "glossary_term",
        "publication",
        "publication_section",
        "framework_core",
        "framework_function",
        "framework_category",
        "framework_subcategory",
        "implementation_example",
        "profile",
        "tier",
        "trustworthiness_characteristic",
        "ai_actor_task",
        "control_family",
        "control",
        "control_enhancement",
        "control_parameter",
        "control_baseline",
        "reference",
        "acronym",
        "mapping",
    ]
    relations = [
        "related_to",
        "contains",
        "part_of",
        "depends_on",
        "supports",
        "implements",
        "implemented_by",
        "calls",
        "called_by",
        "reads_from",
        "writes_to",
        "publishes",
        "subscribes_to",
        "defines",
        "derived_from",
        "supersedes",
        "replaces",
        "owned_by",
        "governed_by",
        "illustrates",
        "helps_achieve",
        "maps_to",
        "has_enhancement",
        "enhances",
        "has_parameter",
        "references",
        "selects_control",
        "included_in_baseline",
        "assessed_by",
        "characterizes",
        "uses",
    ]
    write_text(ROOT / "schemas" / "types.yml", dump_yaml(types) + "\n")
    write_text(ROOT / "schemas" / "relations.yml", dump_yaml(relations) + "\n")


def root_concepts() -> None:
    sources = [
        source_ref("NIST CSF 2.0 PDF", "/source/NIST.CSWP.29.pdf", "pdf", "NIST Cybersecurity Framework 2.0 publication."),
        source_ref("NIST AI RMF 1.0 PDF", "/source/NIST.AI.100-1.pdf", "pdf", "NIST AI RMF 1.0 publication."),
        source_ref("NIST SP 800-53 Rev. 5 PDF", "/source/NIST.SP.800-53r5.pdf", "pdf", "NIST SP 800-53 Rev. 5 publication."),
    ]
    write_text(
        ROOT / "knowledge" / "domains" / "nist-governance-corpus.md",
        md(
            base_fm(
                "domain",
                "NIST Governance Corpus",
                "Cross-publication domain connecting cybersecurity risk, AI risk, and security and privacy controls.",
                ["nist", "governance", "cybersecurity", "ai-risk", "controls"],
                sources,
                [
                    relation("NIST Cybersecurity Framework 2.0", "/bundles/nist-csf-2/knowledge/publication.md", "contains", "The corpus includes CSF 2.0 as a cybersecurity risk management framework."),
                    relation("NIST AI RMF 1.0", "/bundles/nist-ai-rmf-1/knowledge/publication.md", "contains", "The corpus includes AI RMF 1.0 as an AI risk management framework."),
                    relation("NIST SP 800-53 Rev. 5", "/bundles/nist-sp-800-53-r5/knowledge/publication.md", "contains", "The corpus includes SP 800-53 Rev. 5 as a control catalog."),
                ],
                "nist-governance",
                "nist-reference-library",
            ),
            """# Summary

The NIST Governance Corpus brings together three complementary NIST publications: CSF 2.0 for cybersecurity risk outcomes, AI RMF 1.0 for trustworthy AI risk management, and SP 800-53 Rev. 5 for security and privacy controls.

# Scope

This root concept is intentionally cross-publication. Detailed concepts live in nested bundles so each publication remains independently navigable and maintainable.

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)
[3] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)
""",
        ),
    )
    write_text(
        ROOT / "knowledge" / "systems" / "nist-reference-library.md",
        md(
            base_fm(
                "system",
                "NIST Reference Library",
                "Source system for local NIST PDFs and official NIST machine-readable exports used by the EKF demo.",
                ["nist", "source", "reference-library", "oscal", "csf", "ai-rmf"],
                sources + [
                    source_ref("NIST OSCAL SP 800-53 catalog", "/artifacts/data/NIST_SP-800-53_rev5_catalog.json", "oscal", "Downloaded NIST OSCAL catalog used as structured supplemental data."),
                    source_ref("NIST AI RMF Playbook JSON", "/artifacts/data/ai-rmf-playbook.json", "json", "Downloaded NIST AI RMF Playbook JSON used as structured supplemental data."),
                    source_ref("NIST CSF Reference Tool export", "/artifacts/data/csf-2.0-export.xlsx", "xlsx", "Downloaded CSF Reference Tool export used for implementation examples."),
                ],
                [relation("NIST Governance Corpus", "/knowledge/domains/nist-governance-corpus.md", "supports", "The reference library supplies source material for the corpus.")],
                "nist-governance",
                "nist-reference-library",
            ),
            """# Summary

The NIST Reference Library is the source-system concept for this demo. It includes immutable local PDFs plus generated markdown and official NIST structured exports.

# Source Notes

The generated EKF concepts cite the local PDFs as requested source material. NIST machine-readable exports are preserved under `/artifacts/data/` when they provide more precise structure for controls, implementation examples, or playbook entries.

# Citations

[1] [Source manifest](/source/ekf-source.yml)
""",
        ),
    )
    write_text(
        ROOT / "knowledge" / "processes" / "nist-risk-management-integration.md",
        md(
            base_fm(
                "business_process",
                "NIST Risk Management Integration",
                "Integration concept explaining how CSF outcomes, AI RMF outcomes, and SP 800-53 controls can be traversed together.",
                ["nist", "risk-management", "integration", "framework", "controls"],
                sources,
                [
                    relation("CSF Core", "/bundles/nist-csf-2/knowledge/core/csf-core.md", "uses", "CSF Core outcomes structure cybersecurity risk management."),
                    relation("AI RMF Core", "/bundles/nist-ai-rmf-1/knowledge/core/ai-rmf-core.md", "uses", "AI RMF Core outcomes structure AI risk management."),
                    relation("SP 800-53 Control Catalog", "/bundles/nist-sp-800-53-r5/knowledge/publication.md", "supports", "SP 800-53 controls provide selectable safeguards and assessment-oriented detail."),
                ],
                "nist-governance",
                "nist-reference-library",
            ),
            """# Summary

CSF 2.0, AI RMF 1.0, and SP 800-53 Rev. 5 occupy different layers of risk management. CSF and AI RMF organize desired outcomes and lifecycle practices; SP 800-53 provides a catalog of security and privacy controls that can support implementation and assessment.

# Behavior

The graph uses typed relations to move from publications to framework cores, functions, categories, subcategories, control families, controls, and enhancements. It does not claim one-to-one mappings unless a concept explicitly cites a mapping source.

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)
[3] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)
""",
        ),
    )


def build_csf() -> None:
    bundle = "nist-csf-2"
    create_bundle(
        bundle,
        "NIST Cybersecurity Framework 2.0",
        "Nested EKF bundle for NIST CSWP 29, The NIST Cybersecurity Framework (CSF) 2.0.",
        {
            "core": ("CSF Core", "Functions, Categories, and Subcategories of the CSF 2.0 Core."),
            "functions": ("CSF Functions", "Six highest-level CSF 2.0 Functions."),
            "categories": ("CSF Categories", "CSF 2.0 outcome Categories grouped under Functions."),
            "subcategories": ("CSF Subcategories", "CSF 2.0 outcome Subcategories and implementation examples."),
            "profiles": ("CSF Profiles", "Organizational, current, target, and community profile concepts."),
            "tiers": ("CSF Tiers", "Four CSF Tiers for risk governance and risk management rigor."),
            "glossary": ("CSF Glossary", "Shared CSF terminology extracted from the publication."),
        },
    )
    data_dir = ROOT / "artifacts" / "data"
    download(URLS["csf_excel"], data_dir / "csf-2.0-export.xlsx")
    for key in ("csf_functions", "csf_categories", "csf_subcategories", "csf_examples"):
        download(URLS[key], data_dir / f"{key.replace('csf_', 'csf-2.0-')}.json")

    function_defs = {
        "GV": ("GOVERN", "The organization's cybersecurity risk management strategy, expectations, and policy are established, communicated, and monitored"),
        "ID": ("IDENTIFY", "The organization's current cybersecurity risks are understood"),
        "PR": ("PROTECT", "Safeguards to manage the organization's cybersecurity risks are used"),
        "DE": ("DETECT", "Possible cybersecurity attacks and compromises are found and analyzed"),
        "RS": ("RESPOND", "Actions regarding a detected cybersecurity incident are taken"),
        "RC": ("RECOVER", "Assets and operations affected by a cybersecurity incident are restored"),
    }
    categories: dict[str, tuple[str, str, str]] = {}
    subcats: dict[str, tuple[str, str]] = {}
    examples: dict[str, list[str]] = defaultdict(list)
    active_subcategory_ids = extract_csf_active_subcategory_ids()
    wb = load_workbook(data_dir / "csf-2.0-export.xlsx", read_only=True, data_only=True)
    ws = wb["CSF 2.0"]
    current_function = ""
    current_category = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        func, cat, sub, impl, _refs = row[:5]
        if isinstance(func, str) and "(" in func:
            match = re.match(r"(.+?) \(([A-Z]{2})\): (.+)", func)
            if match:
                current_function = match.group(2)
                function_defs[current_function] = (match.group(1), match.group(3))
        if isinstance(cat, str) and "(" in cat and not cat.startswith("[Withdrawn"):
            match = re.match(r"(.+?) \(([A-Z]{2}\.[A-Z]{2})\): (.+)", cat)
            if match:
                current_category = match.group(2)
                if current_category in CSF_ACTIVE_CATEGORIES:
                    categories[current_category] = (current_function, match.group(1), match.group(3))
        if isinstance(sub, str) and ":" in sub and not sub.startswith("[Withdrawn"):
            sid, desc = sub.split(":", 1)
            sid = sid.strip()
            if (
                re.match(r"^[A-Z]{2}\.[A-Z]{2}-\d{2}$", sid)
                and current_category in CSF_ACTIVE_CATEGORIES
                and sid in active_subcategory_ids
            ):
                subcats[sid] = (current_category, " ".join(desc.split()))
                if isinstance(impl, str) and impl.strip():
                    examples[sid].extend(re.findall(r"Ex\d+:\s*(.*?)(?=\nEx\d+:|\Z)", impl.strip(), re.S))

    pub_path = concept_path(bundle, "publication.md")
    core_path = concept_path(bundle, "core/csf-core.md")
    sources = [
        source_ref("NIST CSF 2.0 PDF", "/source/NIST.CSWP.29.pdf", "pdf", "Local source PDF for CSF 2.0."),
        source_ref("CSF Reference Tool export", "/artifacts/data/csf-2.0-export.xlsx", "xlsx", "NIST CSF Reference Tool export used for Categories, Subcategories, and Implementation Examples."),
    ]
    write_text(
        local_concept_path(bundle, "publication.md"),
        md(
            base_fm(
                "publication",
                "NIST CSWP 29 - The NIST Cybersecurity Framework (CSF) 2.0",
                "NIST publication defining CSF 2.0 for managing and communicating cybersecurity risk.",
                ["nist", "csf", "cybersecurity", "risk-management", "framework"],
                sources,
                [relation("CSF Core", core_path, "defines", "The publication defines the CSF Core and explains Profiles, Tiers, and online resources.")],
                "cybersecurity-risk-management",
                "nist-csf-2",
                extra={"artifacts": [{"name": "CSF PDF markdown extraction", "path": PDFS["csf"]["artifact"], "kind": "markdown", "description": "Full page-marked markdown extraction of NIST.CSWP.29.pdf."}]},
            ),
            """# Summary

NIST CSWP 29 describes CSF 2.0, a voluntary framework for helping organizations manage cybersecurity risks through common outcomes, profiles, tiers, and supporting online resources.

# Scope

This bundle models the publication, the CSF Core, all active Functions, Categories, Subcategories, CSF Profiles, CSF Tiers, and implementation examples exported from the NIST CSF Reference Tool.

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [CSF Reference Tool export](/artifacts/data/csf-2.0-export.xlsx)
""",
        ),
    )
    write_text(
        local_concept_path(bundle, "core/csf-core.md"),
        md(
            base_fm(
                "framework_core",
                "CSF Core",
                "The CSF 2.0 Core organizes cybersecurity outcomes into Functions, Categories, and Subcategories.",
                ["nist", "csf", "core", "cybersecurity-outcomes"],
                sources,
                [relation("NIST CSWP 29", pub_path, "part_of", "The CSF Core is defined by the CSF 2.0 publication.")]
                + [relation(title, concept_path(bundle, f"functions/{code.lower()}.md"), "contains", f"The Core contains the {title} Function.") for code, (title, _desc) in function_defs.items()],
                "cybersecurity-risk-management",
                "nist-csf-2",
            ),
            f"""# Summary

The CSF Core provides a taxonomy of cybersecurity outcomes. This generated concept links to {len(function_defs)} Functions, {len(categories)} active Categories, and {len(subcats)} active Subcategories.

# Scope

The Core is outcome-oriented and does not prescribe how outcomes must be achieved. Implementation Examples are embedded in Subcategory concepts when provided by the NIST CSF Reference Tool export.

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [CSF Reference Tool export](/artifacts/data/csf-2.0-export.xlsx)
""",
        ),
    )
    for code, (title, desc) in function_defs.items():
        child_cats = sorted([cid for cid, (f, _t, _d) in categories.items() if f == code])
        write_text(
            local_concept_path(bundle, f"functions/{code.lower()}.md"),
            md(
                base_fm(
                    "framework_function",
                    f"{code} - {title.title()}",
                    desc,
                    ["nist", "csf", "function", code.lower()],
                    sources,
                    [relation("CSF Core", core_path, "part_of", "This Function is part of the CSF Core.")]
                    + [relation(categories[c][1], concept_path(bundle, f"categories/{c.lower()}.md"), "contains", f"{code} contains the {c} Category.") for c in child_cats],
                    "cybersecurity-risk-management",
                    "nist-csf-2",
                    extra={"nist_id": code},
                ),
                f"""# Summary

{desc}

# Structure

This Function contains {len(child_cats)} Categories.

## Categories

{chr(10).join(link(f"../categories/{c.lower()}.md", f"{c} - {categories[c][1]}", categories[c][2]) for c in child_cats)}

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
""",
            ),
        )
    for cid, (fcode, title, desc) in sorted(categories.items()):
        child_subs = sorted([sid for sid, (cat, _desc) in subcats.items() if cat == cid])
        write_text(
            local_concept_path(bundle, f"categories/{cid.lower()}.md"),
            md(
                base_fm(
                    "framework_category",
                    f"{cid} - {title}",
                    desc,
                    ["nist", "csf", "category", fcode.lower(), cid.lower().replace(".", "-")],
                    sources,
                    [relation(function_defs[fcode][0], concept_path(bundle, f"functions/{fcode.lower()}.md"), "part_of", f"This Category is part of the {fcode} Function.")]
                    + [relation(s, concept_path(bundle, f"subcategories/{s.lower()}.md"), "contains", f"{cid} contains Subcategory {s}.") for s in child_subs],
                    "cybersecurity-risk-management",
                    "nist-csf-2",
                    extra={"nist_id": cid},
                ),
                f"""# Summary

{desc}

# Structure

This Category is part of `{fcode}` and contains {len(child_subs)} active Subcategories.

## Subcategories

{chr(10).join(link(f"../subcategories/{s.lower()}.md", s, subcats[s][1]) for s in child_subs)}

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [CSF Reference Tool export](/artifacts/data/csf-2.0-export.xlsx)
""",
            ),
        )
    for sid, (cid, desc) in sorted(subcats.items()):
        exs = [" ".join(e.split()) for e in examples.get(sid, [])]
        body_examples = "\n".join(f"- {e}" for e in exs) or "No implementation examples were present in the exported row for this Subcategory."
        write_text(
            local_concept_path(bundle, f"subcategories/{sid.lower()}.md"),
            md(
                base_fm(
                    "framework_subcategory",
                    sid,
                    desc,
                    ["nist", "csf", "subcategory", sid.lower().replace(".", "-")],
                    sources,
                    [relation(categories[cid][1], concept_path(bundle, f"categories/{cid.lower()}.md"), "part_of", f"This Subcategory is part of the {cid} Category.")],
                    "cybersecurity-risk-management",
                    "nist-csf-2",
                    extra={"nist_id": sid, "implementation_example_count": len(exs)},
                ),
                f"""# Summary

{desc}

# Implementation Examples

{body_examples}

# Source Notes

The Subcategory statement is generated from the CSF 2.0 Reference Tool export and cross-checked against the CSF 2.0 PDF appendix structure.

# Citations

[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)
[2] [CSF Reference Tool export](/artifacts/data/csf-2.0-export.xlsx)
""",
            ),
        )
    profiles = {
        "organizational-profile": ("CSF Organizational Profile", "A mechanism for describing current and target cybersecurity outcomes in an organizational context."),
        "current-profile": ("CSF Current Profile", "A description of cybersecurity outcomes an organization is currently achieving or attempting."),
        "target-profile": ("CSF Target Profile", "A description of desired cybersecurity outcomes selected and prioritized by an organization."),
        "community-profile": ("CSF Community Profile", "A profile for a sector, subsector, technology, threat type, or other shared use case."),
    }
    for pid, (title, desc) in profiles.items():
        write_text(local_concept_path(bundle, f"profiles/{pid}.md"), md(base_fm("profile", title, desc, ["nist", "csf", "profile"], sources, [relation("NIST CSWP 29", pub_path, "part_of", "CSF Profiles are described by the CSF 2.0 publication."), relation("CSF Core", core_path, "uses", "Profiles use CSF Core outcomes as their organizing language.")], "cybersecurity-risk-management", "nist-csf-2"), f"# Summary\n\n{desc}\n\n# Citations\n\n[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)\n"))
    tiers = {
        "tier-1-partial": ("Tier 1 - Partial", "Cybersecurity risk governance and management practices are limited or ad hoc."),
        "tier-2-risk-informed": ("Tier 2 - Risk Informed", "Risk management practices are approved by management but may not be established organization-wide."),
        "tier-3-repeatable": ("Tier 3 - Repeatable", "Risk management practices are formally approved, regularly updated, and consistently applied."),
        "tier-4-adaptive": ("Tier 4 - Adaptive", "The organization adapts cybersecurity practices based on previous and current cybersecurity activity and predictive indicators."),
    }
    for tid, (title, desc) in tiers.items():
        write_text(local_concept_path(bundle, f"tiers/{tid}.md"), md(base_fm("tier", title, desc, ["nist", "csf", "tier"], sources, [relation("NIST CSWP 29", pub_path, "part_of", "CSF Tiers are described by the CSF 2.0 publication."), relation("CSF Organizational Profile", concept_path(bundle, "profiles/organizational-profile.md"), "characterizes", "A Tier can characterize the rigor of profile-related risk governance and management practices.")], "cybersecurity-risk-management", "nist-csf-2"), f"# Summary\n\n{desc}\n\n# Citations\n\n[1] [NIST CSF 2.0 PDF](/source/NIST.CSWP.29.pdf)\n"))

    update_bundle_lists(bundle)


AI_FUNCTIONS = {
    "GOVERN": "Cultivate and implement a culture of risk management within organizations developing, deploying, or acquiring AI systems.",
    "MAP": "Establish the context to frame risks related to an AI system.",
    "MEASURE": "Analyze, assess, benchmark, and monitor AI risk and related impacts.",
    "MANAGE": "Allocate risk resources to mapped and measured AI risks on a regular basis.",
}


def build_ai_rmf() -> None:
    bundle = "nist-ai-rmf-1"
    create_bundle(
        bundle,
        "NIST AI RMF 1.0",
        "Nested EKF bundle for NIST AI 100-1, Artificial Intelligence Risk Management Framework (AI RMF 1.0).",
        {
            "core": ("AI RMF Core", "Govern, Map, Measure, and Manage Functions."),
            "functions": ("AI RMF Functions", "Four AI RMF Core Functions."),
            "categories": ("AI RMF Categories", "Outcome Categories under AI RMF Functions."),
            "subcategories": ("AI RMF Subcategories", "AI RMF Subcategories enriched with NIST Playbook material."),
            "profiles": ("AI RMF Profiles", "AI RMF profile concepts."),
            "trustworthiness": ("Trustworthiness Characteristics", "AI trustworthiness characteristics from the publication."),
            "actors": ("AI Actor Tasks", "AI actor task concepts from the AI RMF appendices."),
            "glossary": ("AI RMF Glossary", "AI RMF terms and concepts."),
        },
    )
    data_path = ROOT / "artifacts" / "data" / "ai-rmf-playbook.json"
    download(URLS["ai_playbook_json"], data_path)
    playbook = json.loads(data_path.read_text(encoding="utf-8"))
    by_function: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in playbook:
        func = str(item["type"]).upper()
        by_function[func].append(item)
        by_category[item["category"]].append(item)

    pub_path = concept_path(bundle, "publication.md")
    core_path = concept_path(bundle, "core/ai-rmf-core.md")
    sources = [
        source_ref("NIST AI RMF 1.0 PDF", "/source/NIST.AI.100-1.pdf", "pdf", "Local source PDF for AI RMF 1.0."),
        source_ref("NIST AI RMF Playbook JSON", "/artifacts/data/ai-rmf-playbook.json", "json", "NIST AI RMF Playbook JSON with subcategory-aligned suggestions."),
    ]
    write_text(local_concept_path(bundle, "publication.md"), md(base_fm("publication", "NIST AI 100-1 - Artificial Intelligence Risk Management Framework (AI RMF 1.0)", "NIST publication defining AI RMF 1.0 for managing risks and promoting trustworthy AI.", ["nist", "ai-rmf", "ai-risk", "trustworthy-ai"], sources, [relation("AI RMF Core", core_path, "defines", "The publication defines the AI RMF Core and Profiles.")], "ai-risk-management", "nist-ai-rmf-1", extra={"artifacts": [{"name": "AI RMF PDF markdown extraction", "path": PDFS["ai"]["artifact"], "kind": "markdown", "description": "Full page-marked markdown extraction of NIST.AI.100-1.pdf."}]}), "# Summary\n\nNIST AI 100-1 defines AI RMF 1.0, a voluntary framework for managing AI risks and improving trustworthiness across the AI lifecycle.\n\n# Citations\n\n[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)\n[2] [NIST AI RMF Playbook JSON](/artifacts/data/ai-rmf-playbook.json)\n"))
    write_text(local_concept_path(bundle, "core/ai-rmf-core.md"), md(base_fm("framework_core", "AI RMF Core", "The AI RMF Core organizes AI risk management outcomes into Govern, Map, Measure, and Manage.", ["nist", "ai-rmf", "core", "ai-risk"], sources, [relation("NIST AI 100-1", pub_path, "part_of", "The AI RMF Core is defined by the AI RMF publication.")] + [relation(name.title(), concept_path(bundle, f"functions/{name.lower()}.md"), "contains", f"The Core contains the {name.title()} Function.") for name in AI_FUNCTIONS], "ai-risk-management", "nist-ai-rmf-1"), f"# Summary\n\nThe AI RMF Core is composed of four Functions and {len(playbook)} Subcategories represented in the NIST AI RMF Playbook JSON.\n\n# Citations\n\n[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)\n"))
    for func, desc in AI_FUNCTIONS.items():
        cats = sorted([c for c in by_category if c.startswith(func + "-")])
        write_text(local_concept_path(bundle, f"functions/{func.lower()}.md"), md(base_fm("framework_function", func.title(), desc, ["nist", "ai-rmf", "function", func.lower()], sources, [relation("AI RMF Core", core_path, "part_of", "This Function is part of the AI RMF Core.")] + [relation(c, concept_path(bundle, f"categories/{c.lower()}.md"), "contains", f"{func} contains Category {c}.") for c in cats], "ai-risk-management", "nist-ai-rmf-1", extra={"nist_id": func}), f"# Summary\n\n{desc}\n\n# Structure\n\nThis Function contains {len(cats)} Categories and {len(by_function[func])} Subcategories.\n\n# Citations\n\n[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)\n"))
    for cat, items in sorted(by_category.items()):
        func = cat.split("-")[0]
        title = f"{cat}: {category_summary(items)}"
        desc = category_summary(items)
        sid_links = []
        for item in items:
            sid = ai_id(item["title"])
            sid_links.append(relation(sid, concept_path(bundle, f"subcategories/{slug(sid)}.md"), "contains", f"{cat} contains Subcategory {sid}."))
        subcategory_list = "\n".join(
            link(
                f"../subcategories/{slug(ai_id(item['title']))}.md",
                ai_id(item["title"]),
                item["description"],
            )
            for item in items
        )
        write_text(
            local_concept_path(bundle, f"categories/{cat.lower()}.md"),
            md(
                base_fm(
                    "framework_category",
                    title,
                    desc,
                    ["nist", "ai-rmf", "category", cat.lower()],
                    sources,
                    [
                        relation(
                            func.title(),
                            concept_path(bundle, f"functions/{func.lower()}.md"),
                            "part_of",
                            f"This Category is part of the {func.title()} Function.",
                        )
                    ]
                    + sid_links,
                    "ai-risk-management",
                    "nist-ai-rmf-1",
                    extra={"nist_id": cat},
                ),
                f"""# Summary

{desc}

# Subcategories

{subcategory_list}

# Citations

[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)
[2] [NIST AI RMF Playbook JSON](/artifacts/data/ai-rmf-playbook.json)
""",
            ),
        )
    for item in playbook:
        sid = ai_id(item["title"])
        cat = item["category"]
        func = item["type"].upper()
        topics = [slug(str(t)) for t in item.get("Topic", []) if str(t).strip()]
        actors = [str(a) for a in item.get("AI Actors", []) if str(a).strip()]
        body = f"""# Summary

{item['description']}

# Playbook About

{item.get('section_about') or 'No Playbook about text was provided for this Subcategory.'}

# Suggested Actions

{item.get('section_actions') or 'No Playbook suggested actions were provided for this Subcategory.'}

# Documentation Prompts

{item.get('section_doc') or 'No Playbook documentation prompts were provided for this Subcategory.'}

# References

{item.get('section_ref') or 'No Playbook references were provided for this Subcategory.'}

# Citations

[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)
[2] [NIST AI RMF Playbook JSON](/artifacts/data/ai-rmf-playbook.json)
"""
        write_text(local_concept_path(bundle, f"subcategories/{slug(sid)}.md"), md(base_fm("framework_subcategory", sid, item["description"], ["nist", "ai-rmf", "subcategory", slug(sid), *topics], sources, [relation(cat, concept_path(bundle, f"categories/{cat.lower()}.md"), "part_of", f"This Subcategory is part of AI RMF Category {cat}.")], "ai-risk-management", "nist-ai-rmf-1", extra={"nist_id": sid, "ai_actors": actors, "topics": item.get("Topic", [])}), body))
    characteristics = {
        "valid-and-reliable": ("Valid and Reliable", "AI systems should be valid, reliable, and perform as intended under expected conditions."),
        "safe": ("Safe", "AI systems should not endanger human life, health, property, or the environment when used as intended or under foreseeable misuse."),
        "secure-and-resilient": ("Secure and Resilient", "AI systems should be protected from adversarial exploitation and recover from adverse events."),
        "accountable-and-transparent": ("Accountable and Transparent", "AI actors should have traceable responsibilities and provide appropriate information about AI systems."),
        "explainable-and-interpretable": ("Explainable and Interpretable", "AI systems should provide information that supports understanding of their operation and outputs."),
        "privacy-enhanced": ("Privacy-Enhanced", "AI systems should preserve privacy and manage privacy risks."),
        "fair-with-harmful-bias-managed": ("Fair - with Harmful Bias Managed", "AI systems should manage harmful bias and support equitable outcomes."),
    }
    for cid, (title, desc) in characteristics.items():
        write_text(local_concept_path(bundle, f"trustworthiness/{cid}.md"), md(base_fm("trustworthiness_characteristic", title, desc, ["nist", "ai-rmf", "trustworthiness"], sources, [relation("NIST AI 100-1", pub_path, "part_of", "Trustworthiness characteristics are described in the AI RMF publication.")], "ai-risk-management", "nist-ai-rmf-1"), f"# Summary\n\n{desc}\n\n# Citations\n\n[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)\n"))
    for pid, title in {"ai-rmf-profile": "AI RMF Profile", "current-profile": "AI RMF Current Profile", "target-profile": "AI RMF Target Profile"}.items():
        write_text(local_concept_path(bundle, f"profiles/{pid}.md"), md(base_fm("profile", title, f"{title} adapts AI RMF outcomes to a specific AI context or organizational use case.", ["nist", "ai-rmf", "profile"], sources, [relation("AI RMF Core", core_path, "uses", "AI RMF Profiles use the Core to describe context-specific outcomes.")], "ai-risk-management", "nist-ai-rmf-1"), f"# Summary\n\n{title} adapts AI RMF outcomes to a specific AI context or organizational use case.\n\n# Citations\n\n[1] [NIST AI RMF 1.0 PDF](/source/NIST.AI.100-1.pdf)\n"))
    update_bundle_lists(bundle)


def ai_id(title: str) -> str:
    return re.sub(r"\s+", " ", title.strip()).upper()


def category_summary(items: list[dict[str, Any]]) -> str:
    descriptions = [str(i.get("description", "")).strip() for i in items if str(i.get("description", "")).strip()]
    if not descriptions:
        return "AI RMF Category generated from Playbook entries."
    first = descriptions[0]
    return first if len(first) < 150 else first[:147].rsplit(" ", 1)[0] + "..."


def text_from_parts(parts: list[dict[str, Any]], names: set[str]) -> str:
    chunks: list[str] = []
    for part in parts:
        if part.get("name") in names and part.get("prose"):
            chunks.append(str(part["prose"]))
        chunks.extend(text_from_parts(part.get("parts", []), names).split("\n\n") if part.get("parts") else [])
    return "\n\n".join(c for c in chunks if c).strip()


def prop(control: dict[str, Any], name: str, class_: str | None = None) -> str:
    for p in control.get("props", []):
        if p.get("name") == name and (class_ is None or p.get("class") == class_):
            return str(p.get("value", ""))
    return ""


def flatten_controls(group: dict[str, Any]) -> tuple[list[dict[str, Any]], list[tuple[dict[str, Any], dict[str, Any]]]]:
    bases = group.get("controls", [])
    enhancements: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for base in bases:
        for child in base.get("controls", []):
            enhancements.append((base, child))
    return bases, enhancements


def build_sp80053() -> None:
    bundle = "nist-sp-800-53-r5"
    create_bundle(
        bundle,
        "NIST SP 800-53 Rev. 5",
        "Nested EKF bundle for NIST SP 800-53 Rev. 5 security and privacy controls.",
        {
            "sections": ("Publication Sections", "Chapters, references, glossary, acronyms, and control summary sections."),
            "families": ("Control Families", "Twenty SP 800-53 control families."),
            "controls": ("Controls", "Base SP 800-53 controls generated from the NIST OSCAL catalog."),
            "enhancements": ("Control Enhancements", "Control enhancements generated from the NIST OSCAL catalog."),
            "glossary": ("Glossary", "Glossary and acronym concepts from the publication."),
            "references": ("References", "Reference and supplemental source concepts."),
        },
    )
    data_path = ROOT / "artifacts" / "data" / "NIST_SP-800-53_rev5_catalog.json"
    download(URLS["sp80053_oscal"], data_path)
    catalog = json.loads(data_path.read_text(encoding="utf-8"))["catalog"]
    sources = [
        source_ref("NIST SP 800-53 Rev. 5 PDF", "/source/NIST.SP.800-53r5.pdf", "pdf", "Local source PDF for SP 800-53 Rev. 5."),
        source_ref("NIST OSCAL SP 800-53 catalog", "/artifacts/data/NIST_SP-800-53_rev5_catalog.json", "oscal", "Official NIST OSCAL catalog used as structured supplemental data for controls and enhancements."),
    ]
    pub_path = concept_path(bundle, "publication.md")
    groups = catalog.get("groups", [])
    base_total = sum(len(g.get("controls", [])) for g in groups)
    enh_total = sum(len(c.get("controls", [])) for g in groups for c in g.get("controls", []))
    write_text(local_concept_path(bundle, "publication.md"), md(base_fm("publication", "NIST SP 800-53 Rev. 5 - Security and Privacy Controls", "NIST publication providing a catalog of security and privacy controls for information systems and organizations.", ["nist", "sp800-53", "controls", "security", "privacy"], sources, [relation(g["title"], concept_path(bundle, f"families/{g['id']}.md"), "contains", f"The publication contains the {g['title']} control family.") for g in groups], "security-and-privacy-controls", "nist-sp-800-53-r5", extra={"artifacts": [{"name": "SP 800-53 PDF markdown extraction", "path": PDFS["sp80053"]["artifact"], "kind": "markdown", "description": "Full page-marked markdown extraction of NIST.SP.800-53r5.pdf."}], "oscal_version": catalog.get("metadata", {}).get("version")}), f"# Summary\n\nNIST SP 800-53 Rev. 5 provides a catalog of security and privacy controls. This EKF bundle represents {len(groups)} control families, {base_total} base controls, and {enh_total} control enhancements from NIST's OSCAL catalog, while preserving the local PDF as source provenance.\n\n# Source Notes\n\nThe local PDF is the requested source publication. The OSCAL catalog is a NIST-maintained machine-readable representation used to generate precise control and enhancement concepts. When the OSCAL release contains later patch-level changes, the concept provenance makes that enrichment explicit.\n\n# Citations\n\n[1] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)\n[2] [NIST OSCAL SP 800-53 catalog](/artifacts/data/NIST_SP-800-53_rev5_catalog.json)\n"))
    section_titles = [
        ("introduction", "Introduction", "Purpose, applicability, audience, responsibilities, relationships, revisions, and publication organization."),
        ("fundamentals", "The Fundamentals", "Requirements and controls, control structure, implementation approaches, and trustworthiness."),
        ("controls", "The Controls", "The SP 800-53 control catalog organized by family."),
        ("references", "References", "Normative and informative references cited by the publication."),
        ("glossary", "Glossary", "Definitions used by the publication."),
        ("acronyms", "Acronyms", "Acronyms used by the publication."),
        ("control-summaries", "Control Summaries", "Summaries of controls and enhancements."),
    ]
    for sid, title, desc in section_titles:
        write_text(local_concept_path(bundle, f"sections/{sid}.md"), md(base_fm("publication_section", title, desc, ["nist", "sp800-53", "section", sid], sources, [relation("NIST SP 800-53 Rev. 5", pub_path, "part_of", "This section is part of the SP 800-53 publication.")], "security-and-privacy-controls", "nist-sp-800-53-r5"), f"# Summary\n\n{desc}\n\n# Citations\n\n[1] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)\n"))
    for group in groups:
        gid = group["id"]
        bases, enhancements = flatten_controls(group)
        family_path = concept_path(bundle, f"families/{gid}.md")
        control_relations = [
            relation(
                f"{control['id'].upper()} - {control['title']}",
                concept_path(bundle, f"controls/{gid}/{control['id']}.md"),
                "contains",
                f"The family contains control {control['id'].upper()}.",
            )
            for control in bases
        ]
        control_list = "\n".join(
            link(
                f"../controls/{gid}/{control['id']}.md",
                f"{control['id'].upper()} - {control['title']}",
                control_summary(control),
            )
            for control in bases
        )
        write_text(
            local_concept_path(bundle, f"families/{gid}.md"),
            md(
                base_fm(
                    "control_family",
                    f"{gid.upper()} - {group['title']}",
                    f"SP 800-53 control family for {group['title']}.",
                    ["nist", "sp800-53", "control-family", gid],
                    sources,
                    [
                        relation(
                            "The Controls",
                            concept_path(bundle, "sections/controls.md"),
                            "part_of",
                            "This family is part of the controls section.",
                        )
                    ]
                    + control_relations,
                    "security-and-privacy-controls",
                    "nist-sp-800-53-r5",
                    extra={
                        "nist_id": gid.upper(),
                        "base_control_count": len(bases),
                        "control_enhancement_count": len(enhancements),
                    },
                ),
                f"""# Summary

The {group['title']} family contains {len(bases)} base controls and {len(enhancements)} control enhancements in the NIST OSCAL catalog.

# Controls

{control_list}

# Citations

[1] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)
[2] [NIST OSCAL SP 800-53 catalog](/artifacts/data/NIST_SP-800-53_rev5_catalog.json)
""",
            ),
        )
        write_text(local_concept_path(bundle, f"controls/{gid}/index.md"), md(index_fm(f"{gid.upper()} Controls", f"Base controls in the {group['title']} family."), f"# {gid.upper()} Controls\n\nBase controls in the {group['title']} family.\n"))
        write_text(local_concept_path(bundle, f"enhancements/{gid}/index.md"), md(index_fm(f"{gid.upper()} Control Enhancements", f"Control enhancements in the {group['title']} family."), f"# {gid.upper()} Control Enhancements\n\nControl enhancements in the {group['title']} family.\n"))
        for control in bases:
            cid = control["id"]
            child_rels = [relation(f"{child['id'].upper()} - {child['title']}", concept_path(bundle, f"enhancements/{gid}/{child['id']}.md"), "has_enhancement", f"{cid.upper()} has enhancement {child['id'].upper()}.") for child in control.get("controls", [])]
            write_text(local_concept_path(bundle, f"controls/{gid}/{cid}.md"), control_doc(bundle, sources, family_path, control, "control", child_rels))
            for child in control.get("controls", []):
                write_text(local_concept_path(bundle, f"enhancements/{gid}/{child['id']}.md"), control_doc(bundle, sources, concept_path(bundle, f"controls/{gid}/{cid}.md"), child, "control_enhancement", [relation(f"{cid.upper()} - {control['title']}", concept_path(bundle, f"controls/{gid}/{cid}.md"), "enhances", f"{child['id'].upper()} enhances base control {cid.upper()}.")]))
    update_bundle_lists(bundle)


def control_summary(control: dict[str, Any]) -> str:
    statement = text_from_parts(control.get("parts", []), {"statement"})
    return " ".join(statement.split())[:220] or f"SP 800-53 control {control['id'].upper()}."


def control_doc(bundle: str, sources: list[dict[str, str]], parent_path: str, control: dict[str, Any], type_: str, rels: list[dict[str, str]]) -> str:
    cid = control["id"]
    title = f"{cid.upper()} - {control['title']}"
    statement = text_from_parts(control.get("parts", []), {"statement"})
    guidance = text_from_parts(control.get("parts", []), {"guidance"})
    params = [p.get("id") for p in control.get("params", []) if p.get("id")]
    links = [l.get("href") or l.get("rel") for l in control.get("links", []) if l.get("href") or l.get("rel")]
    family = cid.split("-")[0]
    parent_name = "Control family" if type_ == "control" else "Base control"
    fm = base_fm(
        type_,
        title,
        control_summary(control),
        ["nist", "sp800-53", type_.replace("_", "-"), family],
        sources,
        [relation(parent_name, parent_path, "part_of", f"{title} is part of its parent SP 800-53 structure.")] + rels,
        "security-and-privacy-controls",
        "nist-sp-800-53-r5",
        extra={
            "nist_id": cid.upper(),
            "sort_id": prop(control, "sort-id"),
            "implementation_level": prop(control, "implementation-level"),
            "parameter_ids": params,
            "reference_links": links,
        },
    )
    return md(
        fm,
        f"""# Summary

{control_summary(control)}

# Statement

{statement or 'No statement text was present in the OSCAL record.'}

# Guidance

{guidance or 'No guidance text was present in the OSCAL record.'}

# Parameters

{chr(10).join(f'- `{p}`' for p in params) if params else 'No parameters were present in the OSCAL record.'}

# Source Notes

This concept was generated from NIST OSCAL structured content and cites the local SP 800-53 PDF as the publication source.

# Citations

[1] [NIST SP 800-53 Rev. 5 PDF](/source/NIST.SP.800-53r5.pdf)
[2] [NIST OSCAL SP 800-53 catalog](/artifacts/data/NIST_SP-800-53_rev5_catalog.json)
""",
    )


def update_bundle_lists(bundle: str) -> None:
    root = ROOT / "bundles" / bundle / "knowledge"
    counts = Counter()
    for path in root.rglob("*.md"):
        if path.name in {"index.md", "log.md"}:
            continue
        text = path.read_text(encoding="utf-8")
        if text.startswith("---"):
            fm = yaml.safe_load(text.split("---", 2)[1]) or {}
            counts[fm.get("type", "unknown")] += 1
    write_text(ROOT / "bundles" / bundle / "artifacts" / "concept-counts.json", json.dumps(counts, indent=2, sort_keys=True) + "\n")


def graph_artifacts() -> None:
    graph_path = ROOT / "artifacts" / "graph" / "graph.json"
    mkdir(graph_path.parent)
    # The validation/parse scripts are run by the shell workflow after this generator.
    template = (SKILL / "assets" / "graph-template.html").read_text(encoding="utf-8")
    if graph_path.exists():
        graph = graph_path.read_text(encoding="utf-8").replace("</", "<\\/")
    else:
        graph = '{"nodes":[],"edges":[],"bundles":[]}'
    write_text(ROOT / "artifacts" / "html" / "knowledge-graph" / "index.html", template.replace("__EKF_GRAPH_JSON__", graph))


def update_meta() -> None:
    write_text(
        ROOT / "meta" / "conventions.md",
        """# NIST Governance EKF Demo Conventions

## Tags

Use lowercase kebab-case tags. Preserve NIST identifiers in `nist_id` frontmatter and in titles.

## Generated Concepts

The generated publication bundles are produced by `scripts/generate_nist_ekf.py`. Manual edits should either update the generator or be documented in `log.md`.

## Search

Prefer `rg` for text discovery and fall back to `grep -R --line-number` when `rg` is unavailable.
""",
    )
    write_text(
        ROOT / "meta" / "governance.md",
        """# NIST Governance EKF Demo Governance

Default owner: Enterprise Knowledge Format Maintainers.

Generated concepts are suitable for demonstration, discovery, and graph traversal. Before operational policy use, compare generated concepts against the cited NIST publication and supplemental NIST data export.
""",
    )
    write_text(
        ROOT / "meta" / "ingestion.md",
        """# NIST Governance EKF Demo Ingestion

This bundle is generated from local PDFs in `/source/` plus official NIST machine-readable exports downloaded into `/artifacts/data/`.

## Workflow

1. Extract every PDF page into `/artifacts/markdown/documents/`.
2. Download CSF Reference Tool export, AI RMF Playbook JSON, and SP 800-53 OSCAL catalog into `/artifacts/data/`.
3. Generate curated EKF concept markdown under root `knowledge/` and publication-specific nested bundles.
4. Validate with `.agents/skills/ekf-bootstrap/scripts/validate_ekf.py`.
5. Parse graph JSON with `.agents/skills/ekf-bootstrap/scripts/parse_ekf_graph.py`.
6. Fill `.agents/skills/ekf-bootstrap/assets/graph-template.html` into `/artifacts/html/knowledge-graph/index.html`.
""",
    )
    write_text(
        ROOT / "log.md",
        """# NIST Governance EKF Demo Log

## 2026-06-22

- Generated complete EKF nested bundles from the three local NIST PDFs.
- Added full page-marked markdown extractions for all source PDFs.
- Added official NIST structured data exports for CSF 2.0, AI RMF Playbook, and SP 800-53 OSCAL.
- Added publication, framework, category, subcategory, control family, control, and control enhancement concepts.
- Added graph JSON and self-contained HTML graph demo artifacts.

## 2026-06-21

- Created EKF starter bundle.
""",
    )


def main() -> None:
    clean_generated()
    root_indexes()
    source_and_artifact_indexes()
    schemas()
    update_meta()
    for key, info in PDFS.items():
        extract_pdf_markdown(key, info)
    root_concepts()
    build_csf()
    build_ai_rmf()
    build_sp80053()
    graph_artifacts()


if __name__ == "__main__":
    main()
